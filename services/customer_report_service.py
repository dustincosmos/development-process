from __future__ import annotations

import math
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from playwright.sync_api import sync_playwright
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from db.development_flow_repository import get_current_product_drawing_for_item
from db.paths import CUSTOMER_FORMS_DIR, CUSTOMER_FORMS_PDF_DIR, TEMPLATES_DIR, find_korean_font
from services.reference_data_service import get_item_row, get_molds


INJECTION_CUSTOMER_TEMPLATE_PATH = TEMPLATES_DIR / "개발_품목_성형조건표_v2_20250817.xlsm"
INJECTION_CUSTOMER_OUTPUT_DIR = CUSTOMER_FORMS_DIR
INJECTION_CUSTOMER_PDF_DIR = CUSTOMER_FORMS_PDF_DIR

INJECTION_STAGE_CELL_MAP = {
    "사출_속도": {10: "D13", 9: "E13", 8: "F13", 7: "G13", 6: "H13", 5: "I13", 4: "J13", 3: "K13", 2: "L13", 1: "M13"},
    "사출_압력": {10: "D14", 9: "E14", 8: "F14", 7: "G14", 6: "H14", 5: "I14", 4: "J14", 3: "K14", 2: "L14", 1: "M14"},
    "사출_거리": {10: "D15", 9: "E15", 8: "F15", 7: "G15", 6: "H15", 5: "I15", 4: "J15", 3: "K15", 2: "L15", 1: "M15"},
    "보압_속도": {3: "D17", 2: "E17", 1: "F17"},
    "보압_압력": {3: "D18", 2: "E18", 1: "F18"},
    "보압_시간": {3: "D19", 2: "E19", 1: "F19"},
    "계량_RPM": {1: "E20", 2: "F20", 3: "G20", 4: "H20"},
    "계량_거리": {1: "E21", 2: "F21", 3: "G21", 4: "H21"},
    "계량_배압": {1: "E22", 2: "F22", 3: "G22", 4: "H22"},
}

INJECTION_EXTRA_CELL_MAP = {
    "쿠션": "G17",
    "석백 전": "D21",
    "석백 후": "I21",
    "실린더_NH": "D25",
    "실린더_N1": "E25",
    "실린더_N2": "F25",
    "실린더_N3": "G25",
    "실린더_N4": "H25",
    "금형온도_고정": "D28",
    "금형온도_이동": "E28",
    "금형이동_특이사항": "F28",
    "금형온도_특이사항": "F28",
    "H/R_번호1": "D31",
    "H/R_번호2": "F31",
    "H/R_번호3": "H31",
    "H/R_번호4": "J31",
    "H/R_온도1": "D32",
    "H/R_온도2": "F32",
    "H/R_온도3": "H32",
    "H/R_온도4": "J32",
    "H/R_특이사항": "L32",
    "사출(충진)_1차": "D35",
    "냉각_1차": "E35",
    "회전_1차": "F35",
    "C/T_1차": "G35",
    "취출방법": "I35",
    "사출(충진)_2차": "D36",
    "냉각_2차": "E36",
    "회전_2차": "F36",
    "C/T_2차": "G36",
}


def build_injection_customer_report_filename(sample_code: str, product_code: str) -> str:
    base_name = f"{sample_code}_{product_code}".strip("_")
    return f"{base_name}.xlsx"


def _resolve_mold_code(sample_row: pd.Series) -> str:
    mold_code = _safe_text(sample_row.get("mold_code"), "")
    if mold_code:
        return mold_code
    item_id = sample_row.get("item_id")
    if item_id is None or (isinstance(item_id, float) and pd.isna(item_id)):
        return ""
    item_row = get_item_row(int(item_id))
    if item_row is None:
        return ""
    return _safe_text(item_row.get("primary_mold_code"), "")


def build_injection_customer_report_pdf_filename(sample_code: str, product_code: str) -> str:
    base_name = f"{sample_code}_{product_code}".strip("_")
    return f"{base_name}.pdf"


def _safe_text(value: object, default: str = "") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def _safe_number(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(numeric):
        return None
    return numeric


def _format_date_text(value: object) -> str:
    text = _safe_text(value)
    if not text:
        return datetime.now().strftime("%Y-%m-%d")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text


def _format_weight_summary(op_detail: dict) -> str:
    weights: list[float] = []
    for idx in range(1, 9):
        numeric = _safe_number(op_detail.get(f"product_weight_{idx}"))
        if numeric is not None:
            weights.append(numeric)
    if not weights:
        return "-"
    return f"{max(weights):.3f}"


def _set_value(ws, cell: str, value: object) -> None:
    ws[cell] = "" if value is None else value


def _trim_to_customer_form(ws) -> None:
    # Customer-facing form is the bordered block in 출력_vba: A3:M59.
    # Remove everything outside that block so the workbook itself only contains the form.
    extra_row_count = max(ws.max_row - 59, 0)
    if extra_row_count:
        ws.delete_rows(60, extra_row_count)
    ws.delete_rows(1, 2)
    extra_col_count = max(ws.max_column - column_index_from_string("M"), 0)
    if extra_col_count:
        ws.delete_cols(column_index_from_string("N"), extra_col_count)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None


def _ensure_korean_font() -> str:
    font_name = "KoreanFont"
    registered = set(pdfmetrics.getRegisteredFontNames())
    if font_name not in registered:
        font_path = find_korean_font()
        if font_path is None:
            return "Helvetica"
        pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
    return font_name


def create_injection_customer_report_pdf(
    *,
    sample_row: pd.Series,
    item_row: pd.Series | None,
    project_row: pd.Series | None,
    experiment_date_text: str,
    experimenter_text: str,
    instruction_detail: dict,
    condition_detail: dict,
    op_detail: dict,
    after_24h_detail: dict,
    second_measurement_detail: dict,
    quality_comment_detail: dict,
    final_comment: str,
    final_action: str,
    current_user_name: str,
) -> Path:
    product_code = _safe_text(item_row.get("product_code") if item_row is not None else "")
    file_name = build_injection_customer_report_pdf_filename(_safe_text(sample_row["sample_code"]), product_code)
    INJECTION_CUSTOMER_PDF_DIR.mkdir(parents=True, exist_ok=True)
    output_path = INJECTION_CUSTOMER_PDF_DIR / file_name

    font_name = _ensure_korean_font()
    page_width, page_height = landscape(A4)
    mb_text = _safe_text(sample_row.get("mb_request_code"), "-")
    mb_ratio = _safe_number(instruction_detail.get("mb_ratio"))
    if mb_ratio is not None:
        mb_text = f"{mb_text} {mb_ratio:.1f}%"
    weight_values = [_safe_number(op_detail.get(f"product_weight_{idx}")) for idx in range(1, 9)]
    weight_values = [v for v in weight_values if v is not None]
    process_source = condition_detail or op_detail
    molds_df = get_molds()
    cavity_display = _safe_text(op_detail.get("cavity"), "-")
    if not molds_df.empty and "used_mold_id" in sample_row.index and pd.notna(sample_row.get("used_mold_id")):
        mold_match = molds_df[molds_df["mold_id"] == int(sample_row["used_mold_id"])]
        if not mold_match.empty and mold_match.iloc[0]["cavity"] not in (None, ""):
            cavity_display = _safe_text(mold_match.iloc[0]["cavity"], "-")
    c = canvas.Canvas(str(output_path), pagesize=landscape(A4))
    c.setTitle("성형 조건표")
    c.setAuthor(current_user_name)

    outer_x = 10 * mm
    outer_y = 8 * mm
    outer_w = page_width - 20 * mm
    outer_h = page_height - 16 * mm
    c.setLineWidth(1.4)
    c.rect(outer_x, outer_y, outer_w, outer_h)

    rows_total = 56.0
    row_h = outer_h / rows_total
    left_w = outer_w * 0.355
    right_w = outer_w - left_w

    def draw_box(x: float, y_top: float, w: float, h: float, line_width: float = 1.0) -> None:
        c.setLineWidth(line_width)
        c.rect(x, y_top - h, w, h)

    def draw_edges(
        x: float,
        y_top: float,
        w: float,
        h: float,
        *,
        left: bool = False,
        right: bool = True,
        top: bool = False,
        bottom: bool = True,
        line_width: float = 0.6,
    ) -> None:
        c.setLineWidth(line_width)
        if left:
            c.line(x, y_top, x, y_top - h)
        if right:
            c.line(x + w, y_top, x + w, y_top - h)
        if top:
            c.line(x, y_top, x + w, y_top)
        if bottom:
            c.line(x, y_top - h, x + w, y_top - h)

    def stage_value(source: dict, prefix: str, stage: int) -> object:
        return source.get(f"{prefix}_{stage}", source.get(f"{prefix}{stage}", ""))

    def draw_text(x: float, y: float, value: str, size: int = 8, bold: bool = False, centered: bool = False) -> None:
        c.setFont(font_name, size)
        if centered:
            c.drawCentredString(x, y, value)
        else:
            c.drawString(x, y, value)

    def wrap_lines(value: str, max_chars: int) -> list[str]:
        text = _safe_text(value, "")
        if not text:
            return []
        lines: list[str] = []
        for raw_line in text.splitlines():
            raw_line = raw_line.strip()
            if not raw_line:
                lines.append("")
                continue
            while len(raw_line) > max_chars:
                lines.append(raw_line[:max_chars])
                raw_line = raw_line[max_chars:]
            lines.append(raw_line)
        return lines

    def draw_multiline(x: float, y_top: float, w: float, h: float, value: str, size: int = 8, pad: float = 3 * mm, line_gap: float = 4.2 * mm) -> None:
        lines = wrap_lines(value, max(10, int((w - 2 * pad) / (size * 0.6))))
        c.setFont(font_name, size)
        cursor = y_top - pad - size
        for line in lines:
            if cursor < y_top - h + pad:
                break
            c.drawString(x + pad, cursor, line)
            cursor -= line_gap

    def cell(x: float, y_top: float, w: float, h: float, label: str = "", value: str = "", label_ratio: float = 0.34) -> None:
        draw_box(x, y_top, w, h, 0.8)
        if label:
            split = x + w * label_ratio
            c.line(split, y_top, split, y_top - h)
            draw_text(x + 2 * mm, y_top - h / 2 + 2, label, 8)
            draw_text(split + 2 * mm, y_top - h / 2 + 2, value or "-", 8)
        else:
            draw_text(x + 2 * mm, y_top - h / 2 + 2, value or "-", 8)

    top = outer_y + outer_h

    # Title
    title_gap_h = row_h * 1.0
    title_h = row_h * 1.0
    current_top = top - title_gap_h
    draw_text(outer_x + outer_w / 2, current_top - title_h / 2 + 4, "성형조건표", 16, centered=True)
    current_top -= title_h
    current_top -= title_gap_h

    # product info top row: experiment date / experimenter
    top_info_h = row_h * 1.0
    top_label_w = outer_w * (72.0 / 1080.0)
    top_value_w = (outer_w - top_label_w * 2) / 2.0
    draw_text(outer_x + 2 * mm, current_top - top_info_h / 2 + 2, "실험일", 8)
    draw_text(outer_x + top_label_w + 2 * mm, current_top - top_info_h / 2 + 2, _safe_text(experiment_date_text, ""), 8)
    draw_text(outer_x + top_label_w + top_value_w + 2 * mm, current_top - top_info_h / 2 + 2, "실험자", 8)
    draw_text(outer_x + top_label_w + top_value_w + top_label_w + 2 * mm, current_top - top_info_h / 2 + 2, _safe_text(experimenter_text or current_user_name, ""), 8)
    current_top -= top_info_h

    # Product info 4 rows
    product_h = row_h * 4.0
    draw_box(outer_x, current_top, outer_w, product_h, 1.0)
    category_w = outer_w * (110.0 / 1080.0)
    c.line(outer_x + category_w, current_top, outer_x + category_w, current_top - product_h)
    draw_text(outer_x + category_w / 2, current_top - product_h / 2 + 3, "제품정보", 10, centered=True)
    mold_code = _resolve_mold_code(sample_row)
    product_rows = [
        [("도번", _safe_text(sample_row.get("base_drawing_revision"), "-"), 4), ("금형제작처", _safe_text(project_row.get("mold_vendor_name") if project_row is not None else "", ""), 2), ("제품중량", f"{max(weight_values):.3f}" if weight_values else "-", 2)],
        [("품명", _safe_text(sample_row.get("item_name"), "-"), 4), ("양산사출처", "선일", 2), ("런너중량", _safe_text(op_detail.get("runner_weight"), "-"), 2)],
        [("포장재개발팀", _safe_text(project_row.get("developer_owner") if project_row is not None else "", ""), 1), ("금형번호", _safe_text(mold_code, "-"), 2), ("감리처", _safe_text(project_row.get("supervisor_name") if project_row is not None else "", ""), 2), ("Cavity", cavity_display, 2)],
        [("톤수", _safe_text(op_detail.get("톤수"), "-"), 1), ("호기", _safe_text(op_detail.get("호기"), "-"), 2), ("사출RESIN", _safe_text(instruction_detail.get("raw_material_label"), "-"), 2), ("MB정보", _safe_text(mb_text, ""), 2)],
    ]
    content_x = outer_x + category_w
    content_w = outer_w - category_w
    product_row_h = product_h / 4.0
    first_w = content_w * (92.0 / 970.0)
    stage_total_w = content_w - first_w
    stage_w_px = stage_total_w / 11.0
    for r_idx, row_values in enumerate(product_rows):
        y_top = current_top - r_idx * product_row_h
        x = content_x
        for idx, (label, value, value_span) in enumerate(row_values):
            if r_idx < 2:
                label_w = first_w if idx == 0 else stage_w_px * 2
            else:
                label_w = first_w if idx in (0, 3) else stage_w_px * 1 if idx in (1, 3) else stage_w_px * 2
                if idx == 2:
                    label_w = stage_w_px * 2
            value_w = stage_w_px * value_span
            cell(x, y_top, label_w + value_w, product_row_h, label, value)
            c.line(x + label_w, y_top, x + label_w, y_top - product_row_h)
            x += label_w + value_w
    current_top -= product_h

    spacer_h = row_h * 2.0
    draw_box(outer_x, current_top, outer_w, spacer_h, 1.0)
    current_top -= spacer_h

    # Experiment 25 rows
    experiment_h = row_h * 25.0
    draw_box(outer_x, current_top, outer_w, experiment_h, 1.0)
    c.line(outer_x + category_w, current_top, outer_x + category_w, current_top - experiment_h)
    draw_text(outer_x + category_w / 2, current_top - experiment_h / 2 + 3, "실험", 10, centered=True)
    exp_x = outer_x + category_w
    exp_w = outer_w - category_w

    def section_box(y_top: float, units: float, title_text: str) -> tuple[float, float]:
        h = row_h * units
        draw_box(exp_x, y_top, exp_w, h, 0.8)
        title_w = exp_w * 0.12
        c.line(exp_x + title_w, y_top, exp_x + title_w, y_top - h)
        draw_text(exp_x + title_w / 2, y_top - h / 2 + 3, title_text, 9, centered=True)
        return exp_x + title_w, h

    exp_top = current_top

    # 사출 4
    sec_h = row_h * 4.0
    draw_box(exp_x, exp_top, exp_w, sec_h, 0.8)
    inj_title_w = exp_w * (78.0 / 970.0)
    inj_label_w = exp_w * (92.0 / 970.0)
    stage_w = (exp_w - inj_title_w - inj_label_w) / 10.0
    rowh = sec_h / 4.0
    draw_box(exp_x, exp_top, inj_title_w, sec_h, 0.6)
    draw_text(exp_x + inj_title_w / 2, exp_top - sec_h / 2 + 3, "사출", 9, centered=True)
    row_labels = ["단계", "속도", "압력", "거리"]
    row_prefixes = ["", "사출_속도", "사출_압력", "사출_거리"]
    for r_idx, row_label in enumerate(row_labels):
        y = exp_top - r_idx * rowh
        draw_box(exp_x + inj_title_w, y, inj_label_w, rowh, 0.6)
        draw_text(exp_x + inj_title_w + 2 * mm, y - rowh / 2 + 2, row_label, 7)
        for stage_idx, stage in enumerate(range(10, 0, -1)):
            x = exp_x + inj_title_w + inj_label_w + stage_idx * stage_w
            draw_box(x, y, stage_w, rowh, 0.6)
            if r_idx == 0:
                draw_text(x + stage_w / 2, y - rowh / 2 + 2, str(stage), 7, centered=True)
            else:
                draw_text(x + stage_w / 2, y - rowh / 2 + 2, _safe_text(stage_value(process_source, row_prefixes[r_idx], stage), ""), 7, centered=True)
    exp_top -= sec_h

    # 보압 4행 7열, 사출 블록 7열 끝 기준
    sec_h = row_h * 4.0
    avail_w = inj_title_w + inj_label_w + stage_w * 5
    rowh = sec_h / 4.0
    widths = [inj_title_w, inj_label_w, stage_w, stage_w, stage_w, stage_w, stage_w]
    row_values = [
        ["단계", "3", "2", "1", "쿠션(mm)", ""],
        ["속도(mm/s)", _safe_text(stage_value(process_source, "보압_속도", 3), ""), _safe_text(stage_value(process_source, "보압_속도", 2), ""), _safe_text(stage_value(process_source, "보압_속도", 1), ""), "", "mm"],
        ["압력(bar)", _safe_text(stage_value(process_source, "보압_압력", 3), ""), _safe_text(stage_value(process_source, "보압_압력", 2), ""), _safe_text(stage_value(process_source, "보압_압력", 1), ""), "", _safe_text(process_source.get("쿠션"), "")],
        ["시간(S)", _safe_text(stage_value(process_source, "보압_시간", 3), ""), _safe_text(stage_value(process_source, "보압_시간", 2), ""), _safe_text(stage_value(process_source, "보압_시간", 1), ""), "", ""],
    ]
    for r_idx in range(4):
        y = exp_top - r_idx * rowh
        if r_idx == 0:
            draw_edges(exp_x, exp_top, widths[0], sec_h, left=True, right=True, top=False, bottom=True, line_width=0.6)
            draw_text(exp_x + widths[0] / 2, exp_top - sec_h / 2 + 2, "보압", 7, centered=True)
        x = exp_x + widths[0]
        for c_idx, w in enumerate(widths[1:]):
            right = True
            bottom = True
            if r_idx == 1 and c_idx == 4:
                right = False
            if r_idx == 1 and c_idx == 4:
                row_values[r_idx][c_idx] = _safe_text(process_source.get("쿠션"), "")
            if r_idx in (2, 3) and c_idx in (4, 5):
                bottom = False
            if r_idx in (2, 3) and c_idx == 5:
                right = False
            draw_edges(x, y, w, rowh, left=False, right=right, top=False, bottom=bottom, line_width=0.6)
            draw_text(x + w / 2, y - rowh / 2 + 2, row_values[r_idx][c_idx], 7, centered=True)
            x += w
    exp_top -= sec_h

    # 계량 3행 8열, 사출 블록 8열 끝 기준
    sec_h = row_h * 3.0
    rowh = sec_h / 3.0
    widths = [inj_title_w, inj_label_w, stage_w, stage_w, stage_w, stage_w, stage_w]
    row_values = [
        ["RPM(%)", "석백전", _safe_text(stage_value(process_source, "계량_RPM", 1), ""), _safe_text(stage_value(process_source, "계량_RPM", 2), ""), _safe_text(stage_value(process_source, "계량_RPM", 3), ""), _safe_text(stage_value(process_source, "계량_RPM", 4), ""), "석백후"],
        ["거리", _safe_text(process_source.get("석백 전"), ""), _safe_text(stage_value(process_source, "계량_거리", 1), ""), _safe_text(stage_value(process_source, "계량_거리", 2), ""), _safe_text(stage_value(process_source, "계량_거리", 3), ""), _safe_text(stage_value(process_source, "계량_거리", 4), ""), _safe_text(process_source.get("석백 후"), "")],
        ["배압", "", _safe_text(stage_value(process_source, "계량_배압", 1), ""), _safe_text(stage_value(process_source, "계량_배압", 2), ""), _safe_text(stage_value(process_source, "계량_배압", 3), ""), _safe_text(stage_value(process_source, "계량_배압", 4), ""), ""],
    ]
    for r_idx in range(3):
        y = exp_top - r_idx * rowh
        if r_idx == 0:
            draw_edges(exp_x, exp_top, widths[0], sec_h, left=True, right=True, top=False, bottom=True, line_width=0.6)
            draw_text(exp_x + widths[0] / 2, exp_top - sec_h / 2 + 2, "계량", 7, centered=True)
        x = exp_x + widths[0]
        for c_idx, w in enumerate(widths[1:]):
            if c_idx in (1, 6) and r_idx == 2:
                x += w
                continue
            if c_idx in (1, 6) and r_idx == 1:
                draw_edges(x, y, w, rowh * 2, left=False, right=True, top=False, bottom=True, line_width=0.6)
                draw_text(x + w / 2, y - rowh + 2, row_values[r_idx][c_idx], 7, centered=True)
                x += w
                continue
            draw_edges(x, y, w, rowh, left=False, right=True, top=False, bottom=True, line_width=0.6)
            draw_text(x + w / 2, y - rowh / 2 + 2, row_values[r_idx][c_idx], 7, centered=True)
            x += w
    exp_top -= sec_h

    # 온도 10, 사출 열 폭 기준
    sec_h = row_h * 10.0
    rowh = sec_h / 10.0
    widths = [inj_title_w, inj_label_w, stage_w, stage_w, stage_w, stage_w, stage_w]
    draw_edges(exp_x, exp_top, widths[0], sec_h, left=True, right=True, top=False, bottom=True, line_width=0.6)
    draw_text(exp_x + widths[0] / 2, exp_top - sec_h / 2 + 2, "온도", 7, centered=True)
    body_x = exp_x + widths[0]
    body_w = inj_label_w + stage_w * 10
    # 실린더 3
    y = exp_top
    cylinder_w = inj_label_w + stage_w * 5
    draw_edges(body_x, y, cylinder_w, rowh, left=False, right=False, top=False, bottom=True, line_width=0.6)
    draw_text(body_x + 2 * mm, y - rowh / 2 + 2, "실린더", 7)
    for rr in range(2):
        yy = y - (rr + 1) * rowh
        x = body_x
        if rr == 0:
            labels = ["", "NH", "N1", "N2", "N3", "N4"]
            temp_widths = [inj_label_w] + [stage_w] * 5
            for ci, w in enumerate(temp_widths):
                draw_edges(x, yy, w, rowh, left=False, right=True, top=False, bottom=True, line_width=0.6)
                draw_text(x + w / 2, yy - rowh / 2 + 2, labels[ci], 7, centered=True)
                x += w
        else:
            values = ["설정", _safe_text(process_source.get("실린더_NH"), ""), _safe_text(process_source.get("실린더_N1"), ""), _safe_text(process_source.get("실린더_N2"), ""), _safe_text(process_source.get("실린더_N3"), ""), _safe_text(process_source.get("실린더_N4"), "")]
            temp_widths = [inj_label_w] + [stage_w] * 5
            for ci, w in enumerate(temp_widths):
                draw_edges(x, yy, w, rowh, left=False, right=True, top=False, bottom=True, line_width=0.6)
                draw_text(x + (2 * mm if ci == 0 else w / 2), yy - rowh / 2 + 2, values[ci], 7, centered=ci != 0)
                x += w
    # 금형온도 3
    y = exp_top - rowh * 3
    draw_edges(body_x, y, body_w, rowh, left=False, right=False, top=False, bottom=True, line_width=0.6)
    draw_text(body_x + 2 * mm, y - rowh / 2 + 2, "금형 온도", 7)
    for rr in range(2):
        yy = y - (rr + 1) * rowh
        x = body_x
        if rr == 0:
            merged = [
                ("", inj_label_w),
                ("고정(섭씨)", stage_w * 2),
                ("이동(섭씨)", stage_w * 2),
                ("특이사항", stage_w * 6),
            ]
            for text_value, w in merged:
                draw_edges(x, yy, w, rowh, left=False, right=True, top=False, bottom=True, line_width=0.6)
                draw_text(x + (2 * mm if text_value == "" else w / 2), yy - rowh / 2 + 2, text_value, 7, centered=text_value != "")
                x += w
        else:
            merged = [
                ("설정", inj_label_w),
                (_safe_text(process_source.get("금형온도_고정"), ""), stage_w * 2),
                (_safe_text(process_source.get("금형온도_이동"), ""), stage_w * 2),
                (_safe_text(process_source.get("금형온도_특이사항") or process_source.get("금형이동_특이사항"), ""), stage_w * 6),
            ]
            for idx, (text_value, w) in enumerate(merged):
                draw_edges(x, yy, w, rowh, left=False, right=True, top=False, bottom=True, line_width=0.6)
                draw_text(x + (2 * mm if idx == 0 else w / 2), yy - rowh / 2 + 2, text_value, 7, centered=idx != 0)
                x += w
    # blank 1 row between mold temp and H/R
    blank_y = exp_top - rowh * 6
    draw_edges(body_x, blank_y, body_w, rowh, left=False, right=False, top=False, bottom=True, line_width=0.6)

    # H/R 3 rows including title
    y = exp_top - rowh * 7
    draw_edges(body_x, y, body_w, rowh, left=False, right=False, top=False, bottom=True, line_width=0.6)
    draw_text(body_x + 2 * mm, y - rowh / 2 + 2, "H/R", 7)
    cols_hr = [inj_label_w, stage_w, stage_w, stage_w, stage_w, stage_w * 6]
    for rr in range(2):
        yy = y - (rr + 1) * rowh
        x = body_x
        if rr == 0:
            values = ["번호", "No1", "No2", "No3", "No4", "특이사항"]
            for ci, w in enumerate(cols_hr):
                draw_edges(x, yy, w, rowh, left=False, right=True, top=False, bottom=True, line_width=0.6)
                draw_text(x + (2 * mm if ci == 0 else w / 2), yy - rowh / 2 + 2, values[ci], 7, centered=ci != 0)
                x += w
        else:
            values = ["온도", _safe_text(process_source.get("H/R_온도1"), ""), _safe_text(process_source.get("H/R_온도2"), ""), _safe_text(process_source.get("H/R_온도3"), ""), _safe_text(process_source.get("H/R_온도4"), ""), _safe_text(process_source.get("H/R_특이사항"), "")]
            for ci, w in enumerate(cols_hr):
                draw_edges(x, yy, w, rowh, left=False, right=True, top=False, bottom=True, line_width=0.6)
                draw_text(x + (2 * mm if ci == 0 or ci == 5 else w / 2), yy - rowh / 2 + 2, values[ci], 7, centered=ci not in (0, 5))
                x += w
    exp_top -= sec_h

    # blank 1 row before cycle time
    exp_top -= row_h

    # cycle time 3
    sec_h = row_h * 3.0
    rowh = sec_h / 3.0
    title_w = widths[0]
    label_w = widths[1]
    stage = widths[2]
    draw_edges(exp_x, exp_top - sec_h, title_w, sec_h, left=True, right=True, top=True, bottom=True, line_width=0.6)
    draw_text(exp_x + 2 * mm, exp_top - sec_h / 2 + 2, "Cycle Time", 7)
    for rr in range(3):
        yy = exp_top - (rr + 1) * rowh
        x = exp_x + title_w
        if rr == 0:
            merged = [
                ("", label_w),
                ("사출(충전)", stage),
                ("냉각", stage),
                ("회전", stage),
                ("C/T", stage),
                ("", stage),
                ("취출방법", stage * 2),
            ]
        elif rr == 1:
            merged = [
                ("1차", label_w),
                (_safe_text(process_source.get("사출(충진)_1차"), ""), stage),
                (_safe_text(process_source.get("냉각_1차"), ""), stage),
                (_safe_text(process_source.get("회전_1차"), ""), stage),
                (_safe_text(process_source.get("C/T_1차"), ""), stage),
                ("", stage),
                (_safe_text(process_source.get("취출방법"), ""), stage * 2),
            ]
        else:
            merged = [
                ("2차", label_w),
                (_safe_text(process_source.get("사출(충진)_2차"), ""), stage),
                (_safe_text(process_source.get("냉각_2차"), ""), stage),
                (_safe_text(process_source.get("회전_2차"), ""), stage),
                (_safe_text(process_source.get("C/T_2차"), ""), stage),
                ("", stage),
                ("", stage * 2),
            ]
        for idx, (text_value, w) in enumerate(merged):
            draw_edges(x, yy, w, rowh, left=False, right=True, top=False, bottom=True, line_width=0.6)
            draw_text(x + (2 * mm if idx in (0, 6) else w / 2), yy - rowh / 2 + 2, text_value, 7, centered=idx not in (0, 6))
            x += w
    current_top -= experiment_h

    # gap 1 row
    gap_h = row_h * 0.0
    current_top -= gap_h

    # lower 4-row 2-column block
    lower_left_w = outer_w * (5.0 / 14.0)
    lower_right_w = outer_w - lower_left_w
    split_x = outer_x + lower_left_w
    issue_title_h = row_h * 1.0
    issue_body_h = row_h * 12.0
    improve_title_h = row_h * 1.0
    improve_body_h = row_h * 7.0

    # issue/measurement block
    draw_box(outer_x, current_top, outer_w, issue_title_h + issue_body_h, 1.0)
    c.line(split_x, current_top, split_x, current_top - issue_title_h - issue_body_h)
    c.line(outer_x, current_top - issue_title_h, outer_x + outer_w, current_top - issue_title_h)
    draw_text(outer_x + 2 * mm, current_top - issue_title_h / 2 + 2, "[문제점 및 현상]", 9)
    draw_text(split_x + 2 * mm, current_top - issue_title_h / 2 + 2, "[중요부 측정 규격]", 9)
    draw_multiline(outer_x, current_top - issue_title_h, lower_left_w, issue_body_h, _safe_text(final_comment, _safe_text(op_detail.get("문제점_현상"), "")), 8)

    # right measurement
    header_h = row_h
    measure_y_top = current_top - issue_title_h
    sample_cols = [12 * mm, 20 * mm, 20 * mm, 20 * mm, 20 * mm, 20 * mm, 20 * mm, 14 * mm, lower_right_w - (12+20*6+14) * mm]
    x = split_x
    headers = ["No", "A 즉시", "A 24H", "B 즉시", "B 24H", "C 즉시", "C 24H", "Cav", "중량"]
    for w, head in zip(sample_cols, headers):
        draw_box(x, measure_y_top, w, header_h, 0.6)
        draw_text(x + w / 2, measure_y_top - header_h / 2 + 2, head, 7, centered=True)
        x += w
    for idx in range(1, 9):
        y = measure_y_top - header_h - (idx - 1) * ((issue_body_h - header_h) / 8.0)
        row_box_h = (issue_body_h - header_h) / 8.0
        x = split_x
        values = [
            str(idx),
            _safe_text(op_detail.get(f"즉시_A_{idx}"), ""),
            _safe_text(after_24h_detail.get(f"24H_A_{idx}"), ""),
            _safe_text(op_detail.get(f"즉시_B_{idx}"), ""),
            _safe_text(after_24h_detail.get(f"24H_B_{idx}"), ""),
            _safe_text(op_detail.get(f"즉시_C_{idx}"), ""),
            _safe_text(after_24h_detail.get(f"24H_C_{idx}"), ""),
            str(idx),
            _safe_text(op_detail.get(f"product_weight_{idx}"), ""),
        ]
        if not any(values[1:7]):
            values = ["", "", "", "", "", "", "", "", ""]
        for w, val in zip(sample_cols, values):
            draw_box(x, y, w, row_box_h, 0.5)
            draw_text(x + w / 2, y - row_box_h / 2 + 2, val, 7, centered=True)
            x += w
    current_top -= issue_title_h + issue_body_h

    # improvement/check block
    draw_box(outer_x, current_top, outer_w, improve_title_h + improve_body_h, 1.0)
    c.line(split_x, current_top, split_x, current_top - improve_title_h - improve_body_h)
    c.line(outer_x, current_top - improve_title_h, outer_x + outer_w, current_top - improve_title_h)
    draw_text(outer_x + 2 * mm, current_top - improve_title_h / 2 + 2, "[개선제안]", 9)
    draw_text(split_x + 2 * mm, current_top - improve_title_h / 2 + 2, "[중요부위 치수 점검 내용]", 9)
    draw_multiline(outer_x, current_top - improve_title_h, lower_left_w, improve_body_h, _safe_text(final_action, _safe_text(op_detail.get("개선사항"), "")), 8)

    point_lines = []
    if second_measurement_detail.get("dimension_checks"):
        point_lines.append("치수/중량 체크: " + ", ".join(second_measurement_detail.get("dimension_checks", [])))
    if second_measurement_detail.get("appearance_checks"):
        point_lines.append("외관 체크: " + ", ".join(second_measurement_detail.get("appearance_checks", [])))
    if second_measurement_detail.get("process_checks"):
        point_lines.append("공정 체크: " + ", ".join(second_measurement_detail.get("process_checks", [])))
    if second_measurement_detail.get("quality_result"):
        point_lines.append("품질 판정: " + _safe_text(second_measurement_detail.get("quality_result"), ""))
    if quality_comment_detail.get("action_checks"):
        point_lines.append("조치 체크: " + ", ".join(quality_comment_detail.get("action_checks", [])))
    if quality_comment_detail.get("issue_summary"):
        point_lines.append("점검 요약: " + _safe_text(quality_comment_detail.get("issue_summary"), ""))
    draw_multiline(split_x, current_top - improve_title_h, lower_right_w, improve_body_h, "\n".join(point_lines), 8)

    c.save()
    return output_path


def create_injection_customer_report_pdf_from_preview_html(
    *,
    preview_html: str,
    file_name: str,
) -> Path:
    INJECTION_CUSTOMER_PDF_DIR.mkdir(parents=True, exist_ok=True)
    output_path = INJECTION_CUSTOMER_PDF_DIR / file_name

    html_document = f"""<!doctype html>
<html lang="ko">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <style>
      html, body {{
        margin: 0;
        padding: 0;
        background: #ffffff;
        font-family: -apple-system, BlinkMacSystemFont, "Apple SD Gothic Neo", "Malgun Gothic", sans-serif;
      }}
      body {{
        padding: 8mm 10mm;
        box-sizing: border-box;
      }}
      .pdf-page {{
        width: 100%;
        box-sizing: border-box;
      }}
    </style>
  </head>
  <body>
    <div class="pdf-page">
      {preview_html}
    </div>
  </body>
</html>
"""

    with tempfile.TemporaryDirectory() as tmp_dir:
        image_path = Path(tmp_dir) / "preview.png"
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch()
            page = browser.new_page(viewport={"width": 1600, "height": 2200}, device_scale_factor=2)
            page.set_content(html_document, wait_until="load")
            page.locator(".pdf-page").screenshot(path=str(image_path))
            browser.close()

        page_width, page_height = landscape(A4)
        margin_x = 10 * mm
        margin_y = 8 * mm
        draw_w = page_width - margin_x * 2
        draw_h = page_height - margin_y * 2
        image = ImageReader(str(image_path))
        img_w, img_h = image.getSize()
        scale = min(draw_w / img_w, draw_h / img_h)
        placed_w = img_w * scale
        placed_h = img_h * scale
        x = margin_x + (draw_w - placed_w) / 2
        y = margin_y + (draw_h - placed_h) / 2

        c = canvas.Canvas(str(output_path), pagesize=landscape(A4))
        c.drawImage(image, x, y, width=placed_w, height=placed_h, preserveAspectRatio=True, mask='auto')
        c.showPage()
        c.save()
    return output_path


def _write_stage_values(ws, op_detail: dict) -> None:
    for prefix, cell_map in INJECTION_STAGE_CELL_MAP.items():
        for stage, cell in cell_map.items():
            _set_value(ws, cell, _safe_text(op_detail.get(f"{prefix}_{stage}")))


def _write_extra_values(ws, op_detail: dict) -> None:
    for field_name, cell in INJECTION_EXTRA_CELL_MAP.items():
        _set_value(ws, cell, _safe_text(op_detail.get(field_name)))


def _write_measurement_values(ws, instruction_detail: dict, op_detail: dict, after_24h_detail: dict) -> None:
    for slot, title_cell, spec_cell in (
        ("A", "F40", "F41"),
        ("B", "H40", "H41"),
        ("C", "J40", "J41"),
    ):
        _set_value(
            ws,
            title_cell,
            _safe_text(instruction_detail.get(f"measurement_title_{slot}") or op_detail.get(f"즉시_{slot}_측정부위"), ""),
        )
        _set_value(ws, spec_cell, _safe_text(instruction_detail.get(f"measurement_spec_{slot}"), ""))

    _set_value(ws, "L41", _safe_text(op_detail.get("runner_weight"), "-"))

    for row_idx in range(1, 9):
        excel_row = 42 + row_idx
        _set_value(ws, f"F{excel_row}", _safe_text(op_detail.get(f"즉시_A_{row_idx}")))
        _set_value(ws, f"G{excel_row}", _safe_text(after_24h_detail.get(f"24H_A_{row_idx}")))
        _set_value(ws, f"H{excel_row}", _safe_text(op_detail.get(f"즉시_B_{row_idx}")))
        _set_value(ws, f"I{excel_row}", _safe_text(after_24h_detail.get(f"24H_B_{row_idx}")))
        _set_value(ws, f"J{excel_row}", _safe_text(op_detail.get(f"즉시_C_{row_idx}")))
        _set_value(ws, f"K{excel_row}", _safe_text(after_24h_detail.get(f"24H_C_{row_idx}")))
        _set_value(ws, f"L{excel_row}", row_idx)
        _set_value(ws, f"M{excel_row}", _safe_text(op_detail.get(f"product_weight_{row_idx}"), "-"))


def create_injection_customer_report(
    *,
    sample_row: pd.Series,
    item_row: pd.Series | None,
    instruction_detail: dict,
    op_detail: dict,
    after_24h_detail: dict,
    final_comment: str,
    final_action: str,
    current_user_name: str,
) -> Path:
    if not INJECTION_CUSTOMER_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"템플릿 파일이 없습니다: {INJECTION_CUSTOMER_TEMPLATE_PATH}")

    product_code = _safe_text(item_row.get("product_code") if item_row is not None else "")
    file_name = build_injection_customer_report_filename(_safe_text(sample_row["sample_code"]), product_code)
    INJECTION_CUSTOMER_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = INJECTION_CUSTOMER_OUTPUT_DIR / file_name

    workbook = load_workbook(INJECTION_CUSTOMER_TEMPLATE_PATH, keep_vba=False)
    worksheet = workbook["출력_vba"]
    for sheet_name in list(workbook.sheetnames):
        if sheet_name != "출력_vba":
            workbook.remove(workbook[sheet_name])
    worksheet.title = "성형조건표"
    worksheet.sheet_state = "visible"

    drawing_info = get_current_product_drawing_for_item(int(sample_row["item_id"]))
    drawing_no = _safe_text((drawing_info or {}).get("drawing_no"), _safe_text(sample_row.get("base_drawing_revision"), "-"))
    cavity_value = _safe_text(op_detail.get("cavity"), "-")
    mb_ratio = _safe_number(instruction_detail.get("mb_ratio"))
    mb_code = _safe_text(sample_row.get("mb_request_code"), "-")

    _set_value(worksheet, "A1", _safe_text(sample_row["sample_code"]))
    _set_value(worksheet, "B4", _format_date_text(sample_row.get("customer_delivery_date")))
    _set_value(worksheet, "G4", _safe_text(current_user_name, "-"))
    _set_value(worksheet, "C5", drawing_no or "-")
    _set_value(worksheet, "C6", _safe_text(sample_row.get("item_name"), "-"))
    _set_value(worksheet, "E7", _safe_text(_resolve_mold_code(sample_row), "-"))
    _set_value(worksheet, "C8", _safe_text(op_detail.get("톤수"), ""))
    _set_value(worksheet, "E8", _safe_text(op_detail.get("호기"), ""))
    _set_value(worksheet, "I8", _safe_text(instruction_detail.get("raw_material_label"), "-"))
    _set_value(worksheet, "L5", _format_weight_summary(op_detail))
    _set_value(worksheet, "L6", _safe_text(op_detail.get("runner_weight"), "-"))
    _set_value(worksheet, "L7", f"1X{cavity_value}" if cavity_value not in ("", "-") else "-")
    _set_value(worksheet, "L8", f"{mb_code}  {mb_ratio:.1f}%" if mb_ratio is not None else mb_code)

    _write_stage_values(worksheet, op_detail)
    _write_extra_values(worksheet, op_detail)
    _set_value(worksheet, "A39", _safe_text(final_comment, _safe_text(op_detail.get("문제점_현상"), "-")))
    _set_value(worksheet, "A52", _safe_text(final_action, _safe_text(op_detail.get("개선사항"), "-")))
    _write_measurement_values(worksheet, instruction_detail, op_detail, after_24h_detail)
    _trim_to_customer_form(worksheet)

    workbook.save(output_path)
    return output_path
